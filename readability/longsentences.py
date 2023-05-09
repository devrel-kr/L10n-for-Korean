import openpyxl
import os

dir_path = input("Enter the directory path: ")

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet.cell(row=1, column=1, value="File Name")
worksheet.cell(row=1, column=2, value="Ratio")

row = 2
for file_name in os.listdir(dir_path):
    if file_name.endswith('.txt'):
        file_path = os.path.join(dir_path, file_name)
        with open(file_path, 'r', encoding='utf-8') as f:
            sentences = f.read().split('\n')
        count = 0
        long_count = 0
        for sentence in sentences:
            if len(sentence) > 25:
                long_count += 1
            count += 1
        if count == 0:
            ratio = "No sentences found."
        else:
            ratio = long_count / count
        worksheet.cell(row=row, column=1, value=file_name)
        worksheet.cell(row=row, column=2, value=ratio)
        row += 1

workbook.save("longsentences_ratio.xlsx")
